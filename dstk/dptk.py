"""
Data processing toolkit
A module containing functions for analysis and manipulation of data-sets, mainly using pandas. Includes:
    - decorators
    - audit
    - transforms cleaning
    - descriptive stats
    - transforms analysis
    - inout
"""
import functools
import logging
import os
import re
import textwrap
import time
from concurrent.futures.process import ProcessPoolExecutor
from datetime import timedelta

import holidays
import pyperclip
import scipy
import xlwings as xw

from fuzzywuzzy import process, fuzz
from google.cloud import bigquery
from openpyxl import load_workbook
from pyxlsb import open_workbook
from tqdm import tqdm

from slibtk import slibtk
from dstk.core import *

logger = logging.getLogger(__name__)


def pycharm_df_autocomplete(df: pd.DataFrame, var_nm: str) -> None:
    func = """def df_column_autocompletes():
        \"\"\"serves no purpose other than making column names autocomplete in pycharm\"\"\"\n\t"""
    rows = '\n\t'.join([f"{var_nm}['{col}'] = {var_nm}['{col}']" for col in df.columns.tolist()])
    pyperclip.copy(func + rows)


def haversine_distance(long1: np.ndarray, lat1: np.ndarray, long2: np.ndarray, lat2: np.ndarray):
    """Calculates the haversine distance between 2 sets of GPS coordinates in df
    note: eudlidean distance run about 3x faster"""
    r = 6371  # average radius of Earth in kilometers

    phi1 = np.radians(lat1)
    phi2 = np.radians(lat2)

    delta_phi = np.radians(lat2 - lat1)
    delta_lambda = np.radians(long2 - long1)

    a = np.sin(delta_phi / 2) ** 2 + np.cos(phi1) * np.cos(phi2) * np.sin(delta_lambda / 2) ** 2
    c = 2 * np.arctan2(np.sqrt(a), np.sqrt(1 - a))
    d = (r * c)  # in kilometers

    return d


def make_distance_matrix(long1: np.ndarray, lat1: np.ndarray, long2: np.ndarray, lat2: np.ndarray,
                         lbl1: Optional[Sequence] = None, lbl2: Optional[Sequence] = None) -> Union[
    np.ndarray, pd.DataFrame]:
    """create a distance (kilometers) matrix of every coordinate  in one set to every coordinate in another set"""
    matrix = np.zeros((lat1.shape[0], lat2.shape[0]))
    for idx, (lat, long) in tqdm(enumerate(zip(lat1, long1)), total=lat1.shape[0]):
        dist = haversine_distance(lat, long, lat2, long2)
        matrix[idx, :] = dist
    if (lbl1 is not None) and (lbl2 is not None):
        matrix = pd.DataFrame(matrix, index=lbl1, columns=lbl2)
    return matrix


def euclidean_distance_matrix(df1: pd.DataFrame, df2: pd.DataFrame, labels: str, eastern: str = 'eastern',
                              northern: str = 'northern', null_diagnoals: bool = True) -> pd.DataFrame:
    """return a distance matrix of every perumtation of a sequences of coordicates, returns the
    distance if km."""
    coord1 = df1[[eastern, northern]].to_numpy().tolist()
    coord2 = df2[[eastern, northern]].to_numpy().tolist()
    meters_in_km = 1_000
    mat = scipy.spatial.distance.cdist(coord1, coord2) / meters_in_km
    if null_diagnoals: np.fill_diagonal(mat, np.NaN)
    return pd.DataFrame(mat, index=df1[labels], columns=df2[labels])


# decorators ###########################################################################################################


def pd_logger(head: bool = True, info: bool = False, coverage: bool = False):
    """
    Decorator that logs properties of the pandas object (DataFrame, Series) returned from the decorated function.
    By default returns the head of the object but can also return output of info method.
    Args:
        head: flag to log header
        info: flag to log info method
    """

    def outer_wrapper(func):
        @functools.wraps(func)
        def inner_wrapper(*args, **kwargs):
            result = func(*args, **kwargs)
            if isinstance(result, (pd.DataFrame, pd.Series)):
                msg = f'pandas return from {func.__name__}'
                if head:
                    msg += f'\n{result.head()}'
                if info:
                    msg += f'\n{result.info()}'
                logger.info(msg)
                if coverage:
                    msg += f'\n{_make_df_coverage_msg(result)}'
            else:
                raise TypeError(f'pd_logger expected a pandas object, got {type(result)}')
            return result

        return inner_wrapper

    return outer_wrapper


def _make_df_coverage_msg(df: pd.DataFrame) -> str:
    """show the percentage of entries in each col that are non-missing"""
    msg = 'The coverage of the merged dataset is:\n'
    msg += str(df.notnull().mean())
    return msg


def np_shape_logger(before=True):
    """
    decorator that logs the shape of the first positional {df, array} passed into the decorated function, before and
    after the func is applied, first positional argument and return of wrapped function must have .shape property

    parameters
    ----------
    before: bool
        flag for logging shape of first positional arg
    """

    def outer_wrapper(func):
        @functools.wraps(func)
        def inner_wrapper(*args, **kwargs) -> pd.DataFrame:
            results = func(*args, **kwargs)

            fmt_msg: Callable = lambda when, nrows, ncols: f'{when}:'.ljust(8) + f'nrows={nrows:,d}'.ljust(
                17) + f'ncols={ncols:,d}'
            if not args:
                args = list(kwargs.values())
            nrows, ncols = args[0].shape
            if before:
                logger.info(fmt_msg('before', nrows, ncols))

            try:
                nrows, ncols = results.shape
                logger.info(fmt_msg('after', nrows, ncols))
            except AttributeError:
                logger.info(f'AttributeError: function does not return object with shape method {type(results)}.')

            return results

        return inner_wrapper

    return outer_wrapper


# transforms cleaning ##################################################################################################


def summary(df: pd.DataFrame) -> None:
    """print a summary of DataFrame characteristics to the terminal"""
    pct_zero = df.select_dtypes(['float']).apply(near2zero).to_frame('close2zero')
    dtypes, nunique = df.dtypes.to_frame('dtypes'), df.nunique().to_frame('nunique')
    stats = pd.concat([dtypes, nunique, df.coverage(), pct_zero], 1)
    print(f'dims={df.shape}')
    print(stats)

    cat_vars = df.nunique().loc[lambda x: x < 20].index.tolist()
    for col in cat_vars:
        print(col)
        print(textwrap.indent(df[col].value_counts_pct().to_string(), '\t'))


def pd_log_with_neg(ser: pd.Series) -> pd.Series:
    """log transform series with negative values by adding constant"""
    return np.log(ser + ser.min() + 1)


def concat_with_keys(dfs: Sequence[pd.DataFrame], keys: Sequence[str], title: str = 'keys') -> pd.DataFrame:
    """append a sequence of dataframes together adding a column that can be used to identify the individual
    dataframes"""
    for key, df in zip(keys, dfs):
        df[title] = key
    return pd.concat(dfs)


def set_dtypes(df: pd.DataFrame, ints: OptSeq = None, floats: OptSeq = None, strs: OptSeq = None, dts: OptSeq = None,
               bools: OptSeq = None) -> pd.DataFrame:
    """convenience function to quickly set the dtypes of DataFrame columns returning the transformed DataFrame"""
    if ints: df[ints] = df[ints].astype(int)
    if floats: df[floats] = df[floats].astype(float)
    if dts: df[dts] = df[dts].apply(pd.to_datetime)
    if strs: df[strs] = df[strs].astype(str)
    if bools: df[bools] = df[bools].apply(bool)
    return df


def set_object_dtype(ser: pd.Series) -> pd.Series:
    """infer the dtype of an object column and set accordingly to a dtype"""
    try:
        return pd.to_datetime(ser)
    except Exception:
        try:
            return pd.to_numeric(ser)
        except Exception:
            return ser.astype(str)


def set_object_dtypes(df: pd.DataFrame) -> pd.DataFrame:
    """see: set_object_dtype()"""
    for col in tqdm(df.select_dtypes(object).columns):
        df[col] = set_object_dtype(df[col])
    return df


def object2numeric(df: pd.DataFrame) -> pd.DataFrame:
    """Set object columns to numeric if possible"""
    obj_cols = df.select_dtypes(object).columns
    df[obj_cols] = df[obj_cols].apply(pd.to_numeric, errors='ignore')
    return df


def show_nulls(df: pd.DataFrame, indexes: Optional = None) -> pd.DataFrame:
    """return just the columns and rows of a DataFrame that contain nulls values"""
    if indexes: df = df.set_index(indexes)
    df = df[df.isnull().any(1)]
    return df[df.columns[df.isnull().any()]]


def impute_with_col_mean(df: pd.DataFrame) -> pd.DataFrame:
    for col in tqdm(df.select_dtypes([float, int]).columns):
        df[col] = df[col].fillna(df[col].mean())
    return df


def near2zero(ser, tol: float = 1e-5) -> float:
    return (np.abs(ser) < tol).mean()


def accounting2numeric(ser: pd.Series) -> pd.Series:
    """Transform a series of numbers in accounting format to numeric type panders column. Making sure to convert
    parentheses to negatives, example:
    before: (56)%
    after: -56"""
    return ser.astype(str).str.replace('\(', '-').str.replace('\)?%?', '').apply(pd.to_numeric, errors='coerce')


def move_cols(df: pd.DataFrame, cols: Union[str, List], pos: int) -> pd.DataFrame:
    """move DataFrame columns to a different position"""
    if isinstance(cols, str): cols = [cols]
    for c in cols:
        ser = df.pop(c)
        df.insert(pos, ser.name, ser)
    return df


def rename_col(df: pd.DataFrame, nm, new_nm) -> pd.DataFrame:
    """rename a DataFrame column and return DataFrame"""
    idx = df.columns.tolist().index(nm)
    ser = df.pop(nm)
    df.insert(idx, new_nm, ser)
    return df


def coverage(df: pd.DataFrame) -> pd.DataFrame:
    """return a DataFrame with pct coverage and pct missing of each col"""
    cov = df.count() / df.shape[0]
    missing = 1 - cov
    df = pd.concat([cov, missing], axis=1)
    df.columns = ['coverage', 'missing']
    return df.applymap('{:.1%}'.format)


def cache_cols(func):
    """decorator that saves original columns as an attribute if that attribute does not already exist"""

    def _inner(df, *args, **kwargs):
        if not hasattr(df, 'orig_cols'): df.orig_cols = df.columns
        return func(df, *args, **kwargs)

    return _inner


@cache_cols
def clean_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Transform DataFrame columns to slug format"""
    df.columns = [slibtk.to_slug(c) for c in df.columns]
    return df


@cache_cols
def lower_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Transform DataFrame columns to lower case"""
    df.columns = df.columns.str.lower()
    return df


@cache_cols
def affix_cols(df: pd.DataFrame, affix: str) -> pd.DataFrame:
    """affix a string to the end of each column name in a DataFrame"""
    df.columns = [f'{c}_{affix}' for c in df.columns]
    return df


def reset_cols(df: pd.DataFrame) -> pd.DataFrame:
    """assign columns to the orig_cols attribute"""
    try:
        df.columns = df.orig_cols
    except AttributeError:
        raise AttributeError("DataFrame requires 'orig_cols' attribute to 'reset_cols'")
    return df


def strip_whitespace(df: pd.DataFrame) -> pd.DataFrame:
    """remove all double,trailing and leading whitespaces in col headers and string cols of the DataFrame"""
    for col in df.select_dtypes(include=['object', 'category']).columns:
        df[col] = df[col].str.split().str.join(' ')
    df.columns = [' '.join(nm.split()) for nm in df.columns]
    return df


def all_duplicates(df: pd.DataFrame, col: str) -> pd.DataFrame:
    """return a DataFrame including all rows that correspond to duplicated values in col"""
    c = df[col]
    df = df[c.isin(c[c.duplicated()])]
    return df.sort_values(col)


@np_shape_logger()
def dropna_decorated(df, subset: List = None) -> pd.DataFrame:
    return df.dropna(subset=subset)


@np_shape_logger()
def mergify(left: pd.DataFrame, right: pd.DataFrame, how: str = "inner", on: Any = None, left_on: Any = None,
            right_on: Any = None, left_index: bool = False, right_index: bool = False, *args, **kwargs) -> pd.DataFrame:
    """wrapper for pd.merge that prints out DataFrame shape before and after the merge"""
    # col_to_use = right.columns.difference(left.columns).tolist()
    return left.merge(right, how, on, left_on, right_on, left_index, right_index, *args, **kwargs)


def mask_loop(df: pd.DataFrame, col: str, conds: List[bool], values: List) -> pd.DataFrame:
    """loop through a list of condition value mask paris applying each pair to col, returning the mutated DataFrame"""
    pairs = list(zip(conds, values))
    for cond, val in pairs:
        df.loc[cond, col] = val
    return df


def encode_and_return_mapping(ser: pd.Series) -> Tuple[pd.Series, Dict[int, str]]:
    """
    encode a series of category values as integers returning the integer representation of the series and the mapping
    from integer to category
    Args:
        ser: series of category values can be str or category dtype

    Returns:
        ser_codes: series with categories encoded as integers
        mapping: dict mapping of integer values to category names
    """
    ser = ser.astype('category')
    mapping = dict(zip(ser.cat.codes, ser.cat.categories))
    ser_codes = ser.cat.codes
    return ser_codes, mapping


def float2int2string(ser):
    """convert columns of objects to string while ensuring the string representation of numbers is an integer ie not
    a float"""
    ser.astype(str).str.replace('\.0+', '')


def fuzzy_match(items: List[str], possible_matches: List[str], col_items: str = 'items') -> pd.DataFrame:
    """Return a DataFrame of two fuzzy matched lists with fuzz score where the iterable contains the individual items
    that are searched for in the lookup list"""
    matches = [(i,) + process.extractOne(i, possible_matches, scorer=fuzz.ratio) for i in tqdm(items)]
    return pd.DataFrame(matches, columns=[col_items, 'matches', 'fuzz_ratio'])


def pd_fuzzy_match(df, col, df_poss, col_poss) -> pd.DataFrame:
    """pandas extension of fuzzy_match()"""
    items = df[col].dropna().unique().tolist()
    possible_matches = df_poss[col_poss].dropna().unique().tolist()
    matches = fuzzy_match(items, possible_matches, col_items=col)
    return df.merge(matches, on=col)


# memory optimisation ##################################################################################################

def make_df(nrow: int, ncol: int) -> pd.DataFrame:
    """make toy DataFrame"""
    df = pd.DataFrame(np.random.randint(100_000, 1_000_000, size=(nrow, ncol)),
                      columns=[f'col{i}' for i in range(ncol)]).astype(float)
    return df


def format_performance_test(df: pd.DataFrame, fname: str = 'temp_df', csv: bool = True) -> pd.DataFrame:
    """test the read, write and memory performance of a DataFrame accross different storage formats"""
    results = {}
    formats = [
        ('parquet', pd.DataFrame.to_parquet, pd.read_parquet),
        ('pickle', pd.DataFrame.to_pickle, pd.read_pickle),
        ('feather', pd.DataFrame.to_feather, pd.read_feather),
    ]
    if csv:
        formats.append(
            ('csv', pd.DataFrame.to_csv, pd.read_csv)
        )

    for fmt, to, _ in formats:
        logger.info(f'to: {fmt}')
        start = time.time()
        to(df, Path('.') / f'{fname}.{fmt}')
        results[fmt] = {'to': time.time() - start}

    for fmt, _, read in formats:
        logger.info(f'read: {fmt}')
        start = time.time()
        path = Path('.') / f'{fname}.{fmt}'
        read(path)
        results[fmt]['read'] = time.time() - start
        results[fmt]['memory'] = path.stat().st_size
        path.unlink()

    res = pd.DataFrame(results)
    res.loc['hr_memory'] = res.loc['memory'].apply(slibtk.hr_bytes)
    return res


def make_df_memory_efficient(df: pd.DataFrame, downcast: bool = False, show_usage: bool = True,
                             **kwargs) -> pd.DataFrame:
    """downcast numeric values > set objects to cats > print of memory usage before and after the processing"""
    if show_usage: get_memory_usage(df)
    if downcast:
        df = _downcast_numeric(df)
    df = _set_objs_to_cats(df, **kwargs)
    if show_usage: get_memory_usage(df)
    return df


def get_memory_usage(pd_obj: Union[pd.DataFrame, pd.Series]) -> None:
    """print out the memory usage of a DataFrame or series in mega bytes"""
    if isinstance(pd_obj, pd.DataFrame):
        bytes_ = pd_obj.memory_usage(deep=True).sum()
    else:
        bytes_ = pd_obj.memory_usage(deep=True)
    print(f'{bytes_ / 1024 ** 2:.2f} MB')


def _downcast_numeric(df: pd.DataFrame) -> pd.DataFrame:
    """downcast int and float cols of a DataFrame"""
    int_cols = df.select_dtypes(['int']).columns.tolist()
    float_cols = df.select_dtypes(['float']).columns.tolist()

    df[int_cols] = df[int_cols].apply(pd.to_numeric, downcast='unsigned')
    df[float_cols] = df[float_cols].apply(pd.to_numeric, downcast='float')

    return df


def _set_objs_to_cats(df: pd.DataFrame, threshold: float = 0.5, logger=None) -> pd.DataFrame:
    """sets any col of type object to type categorical if the number of unique values is below the threshold
    (default=0.5)"""
    cols_obj = df.select_dtypes(['object']).columns
    len_index = len(df.index)

    for col in cols_obj:
        try:
            n_unique = len(df[col].unique())
            ratio = n_unique / len_index
            if logger:
                logger.debug(f'{col} ratio: {ratio}')

            if ratio < threshold:
                df[col] = df[col].astype('category')
        except TypeError:
            print(f'Type error: {col}, unhashable type?')

    return df


def _get_dtype_dict(df: pd.DataFrame) -> Dict:
    """return a dict of col header keys and dtype values"""
    cols = df.columns.tolist()
    dtypes = [i.name for i in df.dtypes.values]
    return dict(zip(cols, dtypes))


# transforms analysis ##################################################################################################


def value_counts_pct(ser: pd.Series, dropna=False, fmt=True) -> pd.DataFrame:
    """take in a series and return the value counts, pct breakdown and cumulative breakdown as a pct formatted
    DataFrame"""
    ser_vc = ser.value_counts(dropna=dropna)
    return cum_pcts(ser_vc, fmt)


def cum_pcts(ser: pd.Series, fmt=True) -> pd.DataFrame:
    """adds a pct breakdown and cumulative pct breakdown to a series"""
    ser = ser.sort_values(ascending=False)
    df = pd.concat(
        [ser,
         ser / ser.sum(),
         ser.cumsum() / ser.sum()
         ], axis=1)
    cols = ['total', 'pct', 'cumulative']
    df.columns = cols
    df.index = df.index.astype(str)
    df.loc['Totals'] = df.sum()
    if fmt:
        df = pd.concat(
            [
                df.iloc[:, 0].apply('{:,}'.format),
                df[cols[1:]].applymap('{:.0%}'.format),
            ], axis=1)

    df.iloc[-1, -1] = np.NaN

    return df


def flatten_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Transform multi index columns names to single level of concatenated names"""
    df.columns = [f'{l0}_{l1}' for l0, l1 in df.columns]
    return df


def row_pcts(df: pd.DataFrame, flatten: bool = False) -> pd.DataFrame:
    """append row wise percentage version of the DataFrame under a multi-index"""
    pcts = df.div(df.sum(1), 0)
    totals = df.sum() / df.sum().sum()
    index = pcts / totals
    df = pd.concat([df, pcts, index], axis=1, keys=['point', 'pct', 'index'])
    if flatten:
        df = flatten_cols(df)
    return df


def col_pcts(df: pd.DataFrame, flatten: bool = False) -> pd.DataFrame:
    """append col wise percentage version of the DataFrame under a multi-index"""
    pcts = df.div(df.sum(), 1)
    totals = df.sum(1) / df.sum(1).sum()
    index = pd.DataFrame(pcts.values / totals.values.reshape(-1, 1), columns=pcts.columns, index=pcts.index)
    df = pd.concat([df, pcts, index], axis=1, keys=['point', 'pct', 'index'])
    if flatten:
        df = flatten_cols(df)
    return df.fillna(0)


def add_count(df: pd.DataFrame, col: str) -> pd.DataFrame:
    """add a value counts of col to df and sort from largest to smallest"""
    new_col = f'n_{col}'
    df[new_col] = df.groupby(col)[col].transform('count')
    return df.sort_values(by=new_col, ascending=False)


def apply_to_filtered_df(df: pd.DataFrame, cols: Collection[str], vals: Collection[str], func: Callable, args=(),
                         **kwargs) -> Dict[str, pd.DataFrame]:
    """
    for every col in cols filter df by the corresponding value in vals and apply
    func storing whatever it returns in a dictionary where the key in the val
    Args:
        df: DataFrame
        cols: col names of the cols that will be filtered by their corresponding value
        vals: The 141 values the cols will be filtered by
        func: The function that will be called on the filtered DataFrames
        args: positional args passed to func
        **kwargs: keyword arguments passed to func

    Returns:
        data: dictionary of string DataFrame key pair values
    """
    if len(cols) != len(vals):
        raise TypeError('Func expects cols and vals to be of equal length '
                        'col:{len(cols)!=len(vals)}')
    data = {}
    for col, val in zip(cols, vals):
        df_filt = df[df[col] == val]
        data[val] = func(df_filt, *args, **kwargs)
    return data


def set_multindex_dtype(df: pd.DataFrame, dtype: str = 'str') -> pd.DataFrame:
    """set the dtype of all levels of a multi-index, by default to string"""
    new_index = [df.index.get_level_values(idx).astype(dtype) for idx in range(len(df.index.levels))]
    df.index = new_index
    return df


def indexes2str(df: pd.DataFrame) -> pd.DataFrame:
    """set the dtype of the df index and col to string type"""
    if isinstance(df.index, pd.MultiIndex):
        df = set_multindex_dtype(df)
    else:
        df.index = df.index.astype(str)
    df.columns = df.columns.astype(str)
    return df


def tidy(df: pd.DataFrame, var_nm: str = 'metric', value_nm: str = 'value') -> pd.DataFrame:
    """transform a df from a cross-tab/matrix format into a tidy formatted df. Typically so that aggregated stats can
    be visualised by data viz apis that only accept tidy data"""
    return indexes2str(df).reset_index().melt(id_vars=df.index.name, var_name=var_nm, value_name=value_nm)


def make_index(df: pd.DataFrame, grp_var: str, cat_var: str, cont_var: str) -> pd.DataFrame:
    """
    take tidy DataFrame and make within group pct breakdown column and index for the groups in grp_var
    Args:
        df: tidy DataFrame
        grp_var: col name of groups
        cat_var: col name of categories in breakdown
        cont_var: continuious variable

    Returns:
        df: tidy DataFrame
    """
    df['pct'] = df.groupby(grp_var)[cont_var].transform(lambda x: x / float(x.sum()))
    totals = (df.groupby(cat_var)[cont_var].sum() / df[cont_var].sum()).to_frame('pct_all').reset_index()
    df = df.merge(totals, on=cat_var, how='left')
    df['index'] = df['pct'] / df['pct_all']
    df['variable'] = cat_var
    return df


def make_group_rank(df: pd.DataFrame, col_cat: Union[str, List], col_cont: str, rank_col: str = 'rank') -> Tuple[
    pd.DataFrame, pd.DataFrame]:
    """add a rank columns to a DataFrame which is based on the sum of a groupby"""
    rank = (df.groupby(col_cat)[col_cont].sum()
            .sort_values(ascending=False)
            .reset_index()
            .reset_index()
            .rename({'index': rank_col}, axis=1)
            .drop(col_cont, 1))
    df = df.merge(rank, on=col_cat, how='left').sort_values(rank_col)
    return df, rank


def bin_categories_lt_thresh(df: pd.DataFrame, thresh: float, grp_var: str, cat_var: str, cont_var: str,
                             non_cat_vars: List[str]) -> pd.DataFrame:
    """bin categories below a certian pct contribution of spend"""
    group_on = [grp_var, cat_var]
    spend_by_range_brand = df.groupby(group_on)[cont_var].sum().reset_index()
    spend_by_range_brand = make_index(spend_by_range_brand, grp_var, cat_var, cont_var)[group_on + ['pct']]
    df = df.merge(spend_by_range_brand, on=group_on)
    brands_other = df.query(f'pct < {thresh}').groupby([grp_var] + non_cat_vars).sum().reset_index()
    brands_other[cat_var] = 'other'
    df = pd.concat([df.query(f'pct >= {thresh}'), brands_other])
    return df


def lag_year(ser: pd.Series) -> pd.Series:
    """parse out year and make log of two part string date column: <year>-<qtr>"""
    df = ser.str.split('-', expand=True)
    df[0] = df[0].astype(int) + 1
    lag = df[0].astype(str) + '-' + df[1]
    lag = lag.rename(ser.name)
    return lag


def make_lag(df: pd.DataFrame, values: List, on: List, date: str, func: Optional[Callable] = None) -> pd.DataFrame:
    """add one year lagged version of passed value columns to a DataFrame and return it"""
    lag = df[values + on]
    if func:
        lag[date] = func(lag[date])
    else:
        lag[date] = lag[date] + pd.DateOffset(years=1)
    for col in values:
        lag = lag.rename_col(col, f'{col}_l1')
    return df.merge(lag, on=on, how='left')


def make_lfl(df: pd.DataFrame, values: List) -> pd.DataFrame:
    """add lfl columns for the passed values. called after make_lag()"""
    for v in values:
        df[f'{v}_lfl'] = (df[f'{v}'] - df[f'{v}_l1']) / df[f'{v}_l1']
    return df


def make_did(df: pd.DataFrame, values: List, on: List, grp_col: str, grp: str) -> pd.DataFrame:
    """make difference in difference from lfl variables"""
    ctrl = df.query(f'{grp_col} == "{grp}"')
    lfls = [f'{col}_lfl' for col in values]
    lfls_ctrl = [f'{col}_ctrl' for col in lfls]
    for old, new in zip(lfls, lfls_ctrl):
        ctrl.rename_col(old, new)
    cols = on + lfls_ctrl
    ctrl = ctrl[cols]
    df = df.merge(ctrl, on=on)
    for col in lfls:
        df[col.replace('lfl', 'did')] = df[col] - df[f'{col}_ctrl']
    return df


def make_lag_lfl(df: pd.DataFrame, values: List, on: List, date: str, func: Optional[Callable] = None) -> pd.DataFrame:
    return make_lfl(make_lag(df, values, on, date, func), values).sort_values(date)


def make_incremental(df: pd.DataFrame, values: List, on: List, date: str) -> pd.DataFrame:
    """Add incremental change in absolute and percentage terms within specified groups"""
    df['rank'] = df.groupby(on)[date].rank()
    p1 = df.query('rank == 1')[on + values]
    for v in values:
        p1 = p1.rename_col(f'{v}', f'{v}_p1')
    df = df.merge(p1, on=on)
    for v in values:
        df[f'{v}_diff'] = df[f'{v}'] - df[f'{v}_p1']
        df[f'{v}_pct_diff'] = df[f'{v}_diff'] / df[f'{v}_p1']
    return df


# time series analysis #################################################################################################


def make_uk_holidays(start=1991, end=2040) -> pd.DataFrame:
    """return a DataFrame of all uk holidays (including easter sunday)"""
    uk = holidays.UK(years=list(range(start, end)))
    uk_easter = {dt: nm for dt, nm in uk.items() if 'Easter Monday' in nm}
    uk_easter_sundays = [(dt - timedelta(days=1), 'Easter Sunday') for dt in uk_easter.keys()]
    uk = pd.DataFrame(list(uk.items()) + uk_easter_sundays, columns=['ds', 'holiday'])
    uk['ds'] = pd.to_datetime(uk['ds'])
    return uk


def chg_from_first_period(ts: pd.DataFrame):
    """for a time series add the change (and pct change) from the first period (row) to each subsequent period as
    additional DataFrames appended under a multi-index"""
    chg = ts.sub(ts.iloc[0, :], axis=1)
    pct = chg.div(ts.iloc[0, :], axis=1)
    return pd.concat([ts, chg, pct], keys=['point', 'chg', 'pct'], axis=1)


def missing_dates(ser: pd.Series):
    """if it returns dates they are missing"""
    return pd.date_range(ser.min(), ser.max()).difference(ser)


# formatting ###########################################################################################################

def apply_fmt_to_df(df: pd.DataFrame, fmt: str = '{:,.0f}', cols: Optional[Sequence] = None) -> pd.DataFrame:
    """format selected cols in a DataFrame"""
    cols = cols if cols else df.columns
    df[cols] = df[cols].applymap(fmt.format)
    return df


def sort_by_first_col(df: pd.DataFrame) -> pd.DataFrame:
    """sort DataFrame descending by its first col"""
    return df.sort_values(df.columns[0], ascending=False)


# multi-processing #####################################################################################################

def pandas_multiproc(df: pd.DataFrame, func: Callable, n_chunks: int) -> pd.DataFrame:
    with ProcessPoolExecutor(max_workers=os.cpu_count()) as executor:
        chunks = np.array_split(df, n_chunks)
        result = executor.map(func, chunks)
    return pd.concat(result)


# in out io ############################################################################################################


def pd_read_excel_range(path: Union[Path, str], sheet: str, cell_range: str) -> pd.DataFrame:
    """
    convert a range of an excel file to a pandas DataFrame where the top row in the col names
    Args:
        path: excel file path
        sheet: name of excel sheet
        cell_range: range of cells to be parsed eg 'G15:Q500000'

    Returns:
        df: cells parsed into DataFrame
    """
    path = Path(path)
    wb = load_workbook(filename=path, read_only=True)
    ws = wb[sheet]
    start, end = cell_range.split(':')
    rows = []
    for row in tqdm(ws[start: end]):
        cols = []
        for cell in row:
            cols.append(cell.value)
        rows.append(cols)
    return pd.DataFrame(rows[1:], columns=rows[0])


def to_gbq(df: pd.DataFrame, destination_table: str, if_exists: str = 'replace') -> None:
    """wrapper for pushing DataFrames to bigquery, note duplicate col names are not accepted"""
    client = bigquery.Client()
    # ensuring cols names are complient with bigquery ie no spaces/punc, starts with alpha
    df.columns = [slibtk.prefix_if_first_is_digit(slibtk.to_slug(col)) for col in df.columns]
    df.to_gbq(destination_table=destination_table,
              project_id=os.environ['GCLOUD_PROJECT'],
              if_exists=if_exists,
              location='europe-west2')


def pd_to_feather_multiproc(df: pd.DataFrame, path: Path, fname: str, n_chunks: int = 16) -> None:
    """
    split a DataFrame in to n chunks and save each one as a feather file in a given directory
    """
    path.mkdir(exist_ok=True, parents=True)
    chunks = enumerate(np.array_split(df, n_chunks))
    to_feather: Callable = lambda dframe, i: dframe.to_feather((path / f'{fname}_{i}.fth').as_posix())
    with ProcessPoolExecutor(max_workers=os.cpu_count()) as e:
        for idx, chunk in tqdm(chunks):
            e.map(to_feather, (chunk,), (idx,))


def load_protected_workbook(path: str, sheet_nm, cell_range='A1:Z50000') -> pd.DataFrame:
    """load protected workbook into a DataFrame"""
    wb = xw.Book(path)
    sheet = wb.sheets[sheet_nm]

    df = sheet[cell_range].options(pd.DataFrame, index=False, headers=True).value

    cols = [col for col in df.columns if col is not None]
    df.dropna(how='all', inplace=True)

    return df[cols]


def excel_dict_writer(frames: Dict[str, pd.DataFrame], writer: pd.ExcelWriter, prefix: Optional[str] = None,
                      **kwargs) -> None:
    """take in a dictionary of string-DataFrame keypair values and save it to an ExcelWriter object"""
    for tab_nm, df in frames.items():
        if prefix:
            tab_nm = prefix + tab_nm
        tab_nm = re.sub(r'\W', '', tab_nm)[:30]
        df.to_excel(writer, sheet_name=tab_nm, **kwargs)


def add_excel_tab(df: pd.DataFrame, sheet_name: str, path: str) -> None:
    """add a DataFrame to an existing excel file at path with with the tab set to sheet name"""
    data = pd.read_excel(path, sheet_name=None)
    sheet_name = re.sub(r'[^\w\s]', '', sheet_name)[:30]
    data[sheet_name] = df


def read_xlsb(path: str) -> pd.DataFrame:
    """function to read in a xlsb format file and return it as a DataFrame"""
    data = []
    with open_workbook(path) as wb:
        for sheetname in wb.sheets:
            with wb.get_sheet(sheetname) as sheet:
                for row in sheet.rows():
                    values = [r.v for r in row]  # retrieving content
                    data.append(values)
    return pd.DataFrame(data[1:], columns=data[0])


# tests ################################################################################################################

def test_daily_coverage(ser: pd.Series):
    """assert every day in period should be included"""
    assert pd.date_range(ser.min(), ser.max()).difference(ser).empty, 'missing daily coverage'


def test_not_empty(pd_obj: DfOrSer):
    assert not pd_obj.empty, f'pandas object is empty {pd_obj}'


def test_no_nulls(pd_obj: DfOrSer):
    assert pd_obj.notnull().any()


def test_no_duplicates(pd_obj: DfOrSer):
    assert pd_obj.duplicated().sum() == 0, f'dupllicate values detected'


# adding functionality as methods to pandas objects
pd.Series.value_counts_pct = value_counts_pct
pd.Series.cum_pcts = cum_pcts
pd.Series.pct = lambda ser: ser / ser.sum()

pd.DataFrame.hr_describe = lambda df: df.describe().applymap(functools.partial(slibtk.hr_numbers))

pd.DataFrame.coverage = coverage
pd.DataFrame.rename_col = rename_col
pd.DataFrame.move_cols = move_cols
pd.DataFrame.row_pcts = row_pcts
pd.DataFrame.col_pcts = col_pcts
pd.DataFrame.summary = summary
pd.DataFrame.mergify = mergify
pd.DataFrame.clean_cols = clean_cols
pd.DataFrame.affix_cols = affix_cols
pd.DataFrame.lower_cols = lower_cols
pd.DataFrame.reset_cols = reset_cols
pd.DataFrame.tidy = tidy
pd.DataFrame.flatten_cols = flatten_cols
pd.DataFrame.set_dtypes = set_dtypes
pd.DataFrame.object2numeric = object2numeric
# pd.DataFrame.pd_to_gbq = pd_to_gbq
